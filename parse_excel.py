import sys
from openpyxl import load_workbook

def parse_excel(file_path, sheet_name=None):
    wb = load_workbook(filename=file_path, data_only=True)
    sheets_to_parse = [sheet_name] if sheet_name else wb.sheetnames
    for sheet in sheets_to_parse:
        ws = wb[sheet]
        print(f"Sheet: {sheet}")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
            print(f"Column A: {row[0]}")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=6, values_only=True):
            print(f"Column F: {row[0]}")
        print("-" * 40)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <excel_file.xlsx> [sheet_name]")
        sys.exit(1)
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    parse_excel(sys.argv[1], sheet_name) 